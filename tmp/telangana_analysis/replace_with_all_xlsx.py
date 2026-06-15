import csv
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook


BASE = Path("/Users/anilkumarkolukulapalli/Documents/dashbord")
SOURCE = Path("/Users/anilkumarkolukulapalli/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/4610BEC9-B38C-46D1-9298-F1A44C8D1E64/all.xlsx")
OUT = BASE / "outputs/telangana_analysis/PowerBI_Build_Pack"
ZIP_PATH = BASE / "outputs/telangana_analysis/Telangana_PowerBI_Build_Pack.zip"
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
SOURCE_SHEET = "VAHAN DIESEL HYBRID CARS 09TH J"


def text(value):
    return str(value or "").replace("\xa0", " ").strip()


def num(value):
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", "").strip()))
    except ValueError:
        return 0


def normalize_fuel(value):
    cleaned = text(value).replace("Disel", "Diesel").replace("disel", "Diesel")
    cleaned = " ".join(cleaned.split()).title()
    cleaned = cleaned.replace("Cng", "CNG")
    return "Electric" if cleaned == "Diesel Electrical" else cleaned


def corrected_fuel(raw_maker, raw_fuel):
    maker = normalize_maker(raw_maker)
    fuel = normalize_fuel(raw_fuel)
    ev_only_makers = {"VINFAST"}
    return "Electric" if maker in ev_only_makers else fuel


def normalize_maker(value):
    raw = re.sub(r"\s+", " ", text(value).upper()).strip()
    raw = raw.replace(" -", "-").replace("- ", "-")
    raw = raw.replace("(P)", "PRIVATE")
    aliases = [
        ("MARUTI SUZUKI", "MARUTI SUZUKI"),
        ("TATA PASSENGER ELECTRIC", "TATA PASSENGER ELECTRIC MOBILITY"),
        ("TATA MOTORS PASSENGER", "TATA MOTORS PASSENGER VEHICLES"),
        ("TATA MOTORS", "TATA MOTORS"),
        ("HYUNDAI", "HYUNDAI"),
        ("MAHINDRA ELECTRIC AUTOMOBILE", "MAHINDRA ELECTRIC"),
        ("MAHINDRA ELECTRIC MOBILITY", "MAHINDRA ELECTRIC"),
        ("MAHINDRA & MAHINDRA", "MAHINDRA & MAHINDRA"),
        ("KIA", "KIA"),
        ("KINETIC", "KINETIC"),
        ("TOYOTA", "TOYOTA"),
        ("JSW MG", "MG MOTOR"),
        ("SKODA AUTO VOLKSWAGEN", "SKODA VOLKSWAGEN"),
        ("VOLKSWAGEN AG", "VOLKSWAGEN"),
        ("HONDA CARS", "HONDA CARS"),
        ("HONDA MOTORCYCLE", "HONDA MOTORCYCLE"),
        ("RENAULT NISSAN", "RENAULT NISSAN"),
        ("NISSAN", "NISSAN"),
        ("RENAULT", "RENAULT"),
        ("BMW", "BMW"),
        ("MERCEDES", "MERCEDES-BENZ"),
        ("BYD", "BYD"),
        ("JAGUAR LAND ROVER", "JAGUAR LAND ROVER"),
        ("STELLANTIS", "STELLANTIS"),
        ("VOLVO", "VOLVO"),
        ("VINFAST", "VINFAST"),
        ("ISUZU", "ISUZU"),
        ("AUDI", "AUDI"),
        ("PORSCHE", "PORSCHE"),
        ("LEXUS", "LEXUS"),
        ("TESLA", "TESLA"),
        ("BENTLEY", "BENTLEY"),
        ("FERRARI", "FERRARI"),
        ("ROLLS ROYCE", "ROLLS ROYCE"),
        ("ATHER", "ATHER"),
        ("BAJAJ", "BAJAJ"),
        ("TVS", "TVS"),
        ("HERO MOTOCORP", "HERO MOTOCORP"),
        ("OLA ELECTRIC", "OLA ELECTRIC"),
        ("GREAVES", "GREAVES ELECTRIC"),
        ("REVOLT", "REVOLT"),
        ("RIVER", "RIVER"),
        ("OMEGA SEIKI", "OMEGA SEIKI"),
        ("BAXY", "BAXY"),
        ("KOMAKI", "KOMAKI"),
        ("SIMPLEENERGY", "SIMPLEENERGY"),
        ("ULTRAVIOLETTE", "ULTRAVIOLETTE"),
        ("BATTRE", "BATTRE"),
        ("GOREEN", "GOREEN"),
        ("OPG MOBILITY", "OPG MOBILITY"),
        ("FORCE MOTORS", "FORCE MOTORS"),
        ("SHREE SHYAM", "SHREE SHYAM AGROTECH"),
        ("WUXI MDKA", "WUXI MDKA"),
        ("OTHERS", "OTHERS"),
    ]
    for pattern, normalized in aliases:
        if pattern in raw:
            return normalized
    cleaned = re.sub(
        r"\b(INDIA|LIMITED|LTD|PRIVATE|PVT|MOTORS?|MOTOR|AUTO|AUTOMOBILES?|COMPANY|CORPN|CORPORATION)\b",
        " ",
        raw,
    )
    return re.sub(r"\s+", " ", cleaned).strip(" -") or raw or "UNKNOWN MAKER"


def parse_office(source_name):
    raw = text(source_name)
    without_ext = re.sub(r"\.xlsx$", "", raw, flags=re.I).strip()
    match = re.search(r"\b(TG\d+)\b", without_ext, flags=re.I)
    office_code = match.group(1).upper() if match else "ALL"
    office = re.sub(r"\s*-\s*TG\d+\s*\([^)]*\)\s*$", "", without_ext, flags=re.I)
    office = re.sub(r"\s*\([^)]*\)\s*$", "", office).strip()
    return office or without_ext or "Unknown Office", office_code


def read_rows():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb[SOURCE_SHEET] if SOURCE_SHEET in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [text(value).upper() for value in next(rows)]
    index = {header: i for i, header in enumerate(headers)}

    cleaned = []
    for source_row in rows:
        if not any(text(value) for value in source_row):
            continue
        source_name = text(source_row[index["RTA NAME"]])
        office, office_code = parse_office(source_name)
        month_values = {month: num(source_row[index[month.upper()]]) for month in MONTHS}
        calculated = sum(month_values.values())
        total = num(source_row[index["TOTAL"]])
        variance = total - calculated
        raw_maker = text(source_row[index["MAKER"]])
        cleaned.append(
            {
                "Source Name": source_name,
                "Office": office,
                "Office Code": office_code,
                "Raw Maker": raw_maker,
                "Maker": normalize_maker(raw_maker),
                "Fuel Type": corrected_fuel(raw_maker, source_row[index["FUEL TYPE"]]),
                **month_values,
                "Total": total,
                "Calculated Total": calculated,
                "Total Variance": variance,
                "Total Check": "OK" if variance == 0 else "Mismatch",
            }
        )
    return cleaned


def write_csv(path, fieldnames, rows):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    wide = read_rows()
    wide_columns = [
        "Source Name",
        "Office",
        "Office Code",
        "Raw Maker",
        "Maker",
        "Fuel Type",
        *MONTHS,
        "Total",
        "Calculated Total",
        "Total Variance",
        "Total Check",
    ]
    write_csv(OUT / "TELANGANA_PowerBI_Wide.csv", wide_columns, wide)

    monthly_rows = []
    for row in wide:
        for month_no, month in enumerate(MONTHS, start=1):
            monthly_rows.append(
                {
                    "Source Name": row["Source Name"],
                    "Office": row["Office"],
                    "Office Code": row["Office Code"],
                    "Raw Maker": row["Raw Maker"],
                    "Maker": row["Maker"],
                    "Fuel Type": row["Fuel Type"],
                    "Month": month,
                    "Month No": month_no,
                    "Month Date": f"2026-{month_no:02d}-01",
                    "Registrations": row[month],
                    "Reported Total": row["Total"],
                    "Calculated Total": row["Calculated Total"],
                    "Total Variance": row["Total Variance"],
                    "Total Check": row["Total Check"],
                }
            )
    write_csv(
        OUT / "TELANGANA_PowerBI_Monthly.csv",
        [
            "Source Name",
            "Office",
            "Office Code",
            "Raw Maker",
            "Maker",
            "Fuel Type",
            "Month",
            "Month No",
            "Month Date",
            "Registrations",
            "Reported Total",
            "Calculated Total",
            "Total Variance",
            "Total Check",
        ],
        monthly_rows,
    )
    write_csv(
        OUT / "Month_Dim.csv",
        ["Month", "Month No", "Month Date", "Month Label"],
        [
            {
                "Month": month,
                "Month No": month_no,
                "Month Date": f"2026-{month_no:02d}-01",
                "Month Label": f"{month} 2026",
            }
            for month_no, month in enumerate(MONTHS, start=1)
        ],
    )
    write_csv(OUT / "Maker_Dim.csv", ["Maker"], [{"Maker": value} for value in sorted({row["Maker"] for row in wide})])
    write_csv(
        OUT / "Fuel_Type_Dim.csv",
        ["Fuel Type"],
        [{"Fuel Type": value} for value in sorted({row["Fuel Type"] for row in wide})],
    )
    offices = {
        (row["Office"], row["Office Code"]): {"Office": row["Office"], "Office Code": row["Office Code"]}
        for row in wide
    }
    write_csv(OUT / "Office_Dim.csv", ["Office", "Office Code"], sorted(offices.values(), key=lambda row: (row["Office"], row["Office Code"])))

    with ZIP_PATH.open("wb") as raw_zip:
        pass
    with zipfile.ZipFile(ZIP_PATH, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_path in sorted(OUT.iterdir()):
            if file_path.is_file():
                zf.write(file_path, file_path.name)

    reported = sum(row["Total"] for row in wide)
    calculated = sum(row["Calculated Total"] for row in wide)
    mismatch = sum(1 for row in wide if row["Total Check"] == "Mismatch")
    print(
        {
            "rows": len(wide),
            "reported": reported,
            "calculated": calculated,
            "variance": reported - calculated,
            "mismatch_rows": mismatch,
            "makers": len({row["Maker"] for row in wide}),
            "offices": len({row["Office"] for row in wide}),
            "fuel_types": len({row["Fuel Type"] for row in wide}),
        }
    )


if __name__ == "__main__":
    main()
